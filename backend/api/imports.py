from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from datetime import datetime
import hashlib
import os
import re
import threading
import queue
import pandas as pd
import json
import logging
import csv
import io
import time
from pathlib import Path
from sqlalchemy.exc import IntegrityError
from backend.utils.db import get_db, get_session_local, get_engine, write_audit
from backend.utils import get_path
from backend.services.alert_workbench import compute_alert_content_hash
from backend.services.snapshot_builder import rebuild_candidate_snapshots


router = APIRouter(prefix="/api/imports", tags=["imports"])

IMPORT_PROGRESS_BATCH_SIZE = 100  # update progress every N rows during sheet parsing
DB_LOCK_RETRY_ATTEMPTS = 12
DB_LOCK_RETRY_DELAY_SECONDS = 1

# Script directory for backup paths
SCRIPT_DIR = Path(__file__).resolve().parent.parent.parent

logger = logging.getLogger(__name__)

# Import queue: files are queued and processed one at a time
_import_queue = queue.Queue()
_queue_worker_started = False

ROW_STATUS_GROUPS = {
    "issues": ("raw_only", "failed"),
    "skipped": ("skipped_duplicate",),
    "raw_only": ("raw_only",),
}
CSV_EXPORT_TYPES = {
    "failures": ("raw_only", "failed"),
    "skipped": ("skipped_duplicate",),
    "raw_only": ("raw_only",),
}


FIELD_ALIASES = {
    "device_id": ["设备ID", "设备 ID", "终端ID", "主机ID", "device_id", "Device ID"],
    "first_alert_time": ["首次告警时间", "首次告警", "first_alert_time", "First Alert Time"],
    "last_alert_time": ["最近告警时间", "最后告警时间", "告警时间", "last_alert_time", "Last Alert Time"],
    "source_ip": ["源IP", "源 IP", "内网IP", "src_ip", "source_ip", "Source IP"],
    "target": ["外联目标", "目标", "访问目标", "目的IP", "域名", "target", "Target", "dst"],
    "port": ["外联端口", "端口", "目的端口", "dst_port", "port", "Port"],
    "threat_type": ["威胁类型", "threat_type", "Threat Type"],
    "threat_level": ["威胁等级", "threat_level", "Threat Level"],
    "std_apt_org": ["标准APT组织", "标准 APT 组织", "std_apt_org", "standard_apt_org"],
    "apt_org": ["APT组织", "APT 组织", "组织", "apt_org", "APT Org"],
    "apt_org_tier": ["APT组织分类", "APT 分级", "APT组织分级", "apt_org_tier"],
    "alert_count": ["告警次数", "次数", "alert_count", "Alert Count"],
    "vendors": ["厂商", "vendors", "Vendor"],
    "protocol": ["协议", "protocol", "Protocol"],
    "intel_tags": ["情报标签", "标签", "intel_tags", "Intel Tags"],
    "target_type": ["目标类型", "target_type", "Target Type"],
    "dns_resolved_ip": ["DNS解析IP", "DNS 解析 IP", "dns_resolved_ip"],
    "down_traffic": ["下行流量", "down_traffic"],
    "up_traffic": ["上行流量", "up_traffic"],
    "asset_type": ["资产类型", "asset_type"],
    "intel_position": ["情报位置", "intel_position"],
    "disposal_action": ["处置动作", "disposal_action"],
    "analysis_status": ["研判状态", "分析状态", "analysis_status"],
    "is_focused": ["重点关注", "是否关注", "is_focused"],
}


def _norm_column(value):
    return re.sub(r"\s+", "", str(value or "")).lower()


def _resolve_columns(columns):
    normalized = {_norm_column(column): column for column in columns}
    resolved = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            column = normalized.get(_norm_column(alias))
            if column is not None:
                resolved[field] = column
                break
    return resolved


def _json_safe(value):
    if pd.isna(value):
        return None
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime().isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip() if not isinstance(value, (int, float, bool)) else value


def _clean_cell(value):
    if pd.isna(value):
        return None
    text_value = str(value).strip()
    if not text_value or text_value.lower() == "nan":
        return None
    return text_value


def _get_cell(row, columns, field, default=None):
    column = columns.get(field)
    if not column:
        return default
    value = _clean_cell(row[column])
    return default if value is None else value


def _parse_datetime(value):
    text_value = _clean_cell(value)
    if not text_value:
        return None
    parsed = pd.to_datetime(
        text_value.replace("/", "-").replace(".", "-"),
        errors="coerce",
    )
    if pd.isna(parsed):
        return text_value
    return parsed.to_pydatetime().strftime("%Y-%m-%d %H:%M:%S")


def _parse_int(value, default=0):
    text_value = _clean_cell(value)
    if text_value is None:
        return default
    try:
        return int(float(text_value.replace(",", "")))
    except Exception:
        return default


def _parse_bool_flag(value, default=0):
    """Parse boolean-like fields (is_focused) from Excel. Handles 是/否/1/0/True/False."""
    text_value = _clean_cell(value)
    if text_value is None:
        return default
    lowered = text_value.lower().strip()
    if lowered in ("是", "yes", "true", "1", "y", "是"):
        return 1
    if lowered in ("否", "no", "false", "0", "n", ""):
        return 0
    try:
        return int(float(text_value.replace(",", "")))
    except Exception:
        return default


def _row_payload(row):
    values = row.to_dict() if hasattr(row, "to_dict") else dict(row)
    return {str(key): _json_safe(value) for key, value in values.items()}


def _row_hash(source_file, sheet_name, excel_row_number, raw_json):
    raw = f"{source_file}|{sheet_name}|{excel_row_number}|{raw_json}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def _update_import(db, import_id, **fields):
    if not fields:
        return
    params = {"id": import_id}
    updates = []
    for key, value in fields.items():
        updates.append(f"{key} = :{key}")
        params[key] = value
    db.execute(text(f"UPDATE imports SET {', '.join(updates)} WHERE id = :id"), params)


def _resolve_row_status_filters(status=None, status_group=None):
    if status:
        return (status,)
    if status_group:
        return ROW_STATUS_GROUPS.get(status_group, ())
    return ()


def _extract_export_value(raw_payload, normalized_payload, normalized_key, raw_keys):
    normalized_payload = normalized_payload or {}
    raw_payload = raw_payload or {}
    value = normalized_payload.get(normalized_key)
    if value not in (None, ""):
        return value
    for key in raw_keys:
        value = raw_payload.get(key)
        if value not in (None, ""):
            return value
    return None


def _build_failure_export_rows(rows):
    exported = []
    for row in rows:
        raw_payload = row.get("raw")
        if raw_payload is None:
            raw_payload = json.loads(row.get("raw_json") or "{}")
        normalized_payload = row.get("normalized")
        if normalized_payload is None:
            normalized_payload = json.loads(row.get("normalized_json") or "{}") if row.get("normalized_json") else {}
        exported.append({
            "sheet": row.get("sheet_name"),
            "row": row.get("excel_row_number"),
            "status": row.get("parse_status"),
            "error": row.get("parse_error"),
            "device_id": _extract_export_value(raw_payload, normalized_payload, "device_id", ["设备ID", "设备 ID", "device_id"]),
            "target": _extract_export_value(raw_payload, normalized_payload, "target", ["外联目标", "目标", "target", "Target"]),
            "port": _extract_export_value(raw_payload, normalized_payload, "port", ["外联端口", "端口", "port", "Port"]),
        })
    return exported


def _should_run_vacuum(page_size, freelist_count, page_count):
    if not page_size or not freelist_count or not page_count:
        return False
    reclaim_bytes = page_size * freelist_count
    reclaim_ratio = freelist_count / max(page_count, 1)
    return reclaim_bytes >= 32 * 1024 * 1024 and reclaim_ratio >= 0.1


def _checkpoint_after_import_delete(raw_conn):
    raw_conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")


def _load_alert_id_map_for_hashes(db, content_hashes, batch_size=500):
    hash_to_alert_id = {}
    hashes = [item for item in content_hashes if item]
    for start in range(0, len(hashes), batch_size):
        batch = hashes[start:start + batch_size]
        if not batch:
            continue
        placeholders = ", ".join(f":hash_{idx}" for idx in range(len(batch)))
        rows = db.execute(text(
            "SELECT id, content_hash FROM alerts WHERE content_hash IN "
            f"({placeholders})"
        ), {f"hash_{idx}": value for idx, value in enumerate(batch)}).fetchall()
        for row in rows:
            hash_to_alert_id[row[1]] = row[0]
    return hash_to_alert_id


def _rebuild_import_row_updates_for_repair(rows, parsed_alert_map=None, existing_hashes=None):
    parsed_alert_map = parsed_alert_map or {}
    workbook_seen_hashes = set(existing_hashes or set())
    updates = []
    stats = {"parsed": 0, "raw_only": 0, "failed": 0}

    for row in rows:
        raw_payload = json.loads(row.get("raw_json") or "{}")
        try:
            item, parse_error = _alert_from_row(raw_payload, _resolve_columns(list(raw_payload.keys())))
            if parse_error:
                stats["raw_only"] += 1
                updates.append({
                    "id": row["id"],
                    "parse_status": "raw_only",
                    "parse_error": parse_error,
                    "normalized_json": None,
                    "alert_id": None,
                })
                continue

            content_hash = item["content_hash"]
            alert_info = parsed_alert_map.get(row["id"])
            if alert_info:
                stats["parsed"] += 1
                workbook_seen_hashes.add(content_hash)
                updates.append({
                    "id": row["id"],
                    "parse_status": "parsed",
                    "parse_error": None,
                    "normalized_json": json.dumps(item, ensure_ascii=False, default=str),
                    "alert_id": alert_info.get("alert_id"),
                })
                continue

            if content_hash in workbook_seen_hashes:
                stats["parsed"] += 1
                updates.append({
                    "id": row["id"],
                    "parse_status": "skipped_duplicate",
                    "parse_error": None,
                    "normalized_json": json.dumps(item, ensure_ascii=False, default=str),
                    "alert_id": None,
                })
                continue

            workbook_seen_hashes.add(content_hash)
            stats["parsed"] += 1
            updates.append({
                "id": row["id"],
                "parse_status": "parsed",
                "parse_error": None,
                "normalized_json": json.dumps(item, ensure_ascii=False, default=str),
                "alert_id": None,
            })
        except Exception as exc:
            stats["failed"] += 1
            updates.append({
                "id": row["id"],
                "parse_status": "failed",
                "parse_error": str(exc),
                "normalized_json": None,
                "alert_id": None,
            })

    return updates, stats


def _is_db_locked_error(exc):
    return "database is locked" in str(exc).lower()


def _run_locked_retry(db, operation, *, retries=DB_LOCK_RETRY_ATTEMPTS):
    last_error = None
    for attempt in range(retries):
        try:
            return operation()
        except Exception as exc:
            if not _is_db_locked_error(exc) or attempt == retries - 1:
                raise
            last_error = exc
            db.rollback()
            time.sleep(DB_LOCK_RETRY_DELAY_SECONDS)
    raise last_error


def _alert_from_row(row, columns):
    device_id = _get_cell(row, columns, "device_id")
    source_ip = _get_cell(row, columns, "source_ip")
    target = _get_cell(row, columns, "target")
    first_alert_time = _parse_datetime(_get_cell(row, columns, "first_alert_time"))
    last_alert_time = _parse_datetime(_get_cell(row, columns, "last_alert_time"))
    if not first_alert_time:
        first_alert_time = last_alert_time
    if not last_alert_time:
        last_alert_time = first_alert_time

    missing = []
    if not device_id and not source_ip:
        missing.append("设备ID或源IP")
    if not target:
        missing.append("外联目标")
    if not last_alert_time:
        missing.append("告警时间")
    if missing:
        return None, "缺少核心字段: " + "、".join(missing)

    if not device_id:
        device_id = source_ip
    if not source_ip:
        source_ip = device_id

    port = _get_cell(row, columns, "port")
    std_apt_org = _get_cell(row, columns, "std_apt_org")
    item = {
        "device_id": device_id,
        "first_alert_time": first_alert_time,
        "last_alert_time": last_alert_time,
        "source_ip": source_ip,
        "target": target,
        "target_type": _get_cell(row, columns, "target_type"),
        "port": port,
        "threat_type": _get_cell(row, columns, "threat_type"),
        "threat_level": _get_cell(row, columns, "threat_level"),
        "std_apt_org": std_apt_org.lower() if std_apt_org else None,
        "apt_org": _get_cell(row, columns, "apt_org"),
        "apt_org_tier": _get_cell(row, columns, "apt_org_tier"),
        "alert_count": _parse_int(_get_cell(row, columns, "alert_count")),
        "vendors": _get_cell(row, columns, "vendors"),
        "protocol": _get_cell(row, columns, "protocol"),
        "intel_tags": _get_cell(row, columns, "intel_tags"),
        "intel_position": _get_cell(row, columns, "intel_position"),
        "disposal_action": _get_cell(row, columns, "disposal_action"),
        "analysis_status": _get_cell(row, columns, "analysis_status"),
        "is_focused": _parse_bool_flag(_get_cell(row, columns, "is_focused")),
        "dns_resolved_ip": _get_cell(row, columns, "dns_resolved_ip"),
        "down_traffic": _parse_int(_get_cell(row, columns, "down_traffic")),
        "up_traffic": _parse_int(_get_cell(row, columns, "up_traffic")),
        "asset_type": _get_cell(row, columns, "asset_type"),
    }
    item["content_hash"] = compute_alert_content_hash(item)
    item["unique_hash"] = item["content_hash"]
    return item, None


def _insert_alert(db, item):
    row = db.execute(text(
        "SELECT id FROM alerts WHERE content_hash = :content_hash LIMIT 1"
    ), {"content_hash": item["content_hash"]}).fetchone()
    if row:
        return False, row[0]

    try:
        result = db.execute(text(
        "INSERT INTO alerts ("
        "device_id, first_alert_time, last_alert_time, source_ip, target, "
        "target_type, port, threat_type, threat_level, std_apt_org, "
        "apt_org, apt_org_tier, alert_count, vendors, protocol, "
        "intel_tags, intel_position, disposal_action, analysis_status, is_focused, "
        "dns_resolved_ip, down_traffic, up_traffic, asset_type, "
        "source_file, imported_at, unique_hash, content_hash, import_id, import_sheet_id, "
        "import_row_id, sheet_name, excel_row_number, raw_row_hash"
        ") VALUES ("
        ":device_id, :first_alert_time, :last_alert_time, :source_ip, :target, "
        ":target_type, :port, :threat_type, :threat_level, :std_apt_org, "
        ":apt_org, :apt_org_tier, :alert_count, :vendors, :protocol, "
        ":intel_tags, :intel_position, :disposal_action, :analysis_status, :is_focused, "
        ":dns_resolved_ip, :down_traffic, :up_traffic, :asset_type, "
        ":source_file, :imported_at, :unique_hash, :content_hash, :import_id, :import_sheet_id, "
        ":import_row_id, :sheet_name, :excel_row_number, :raw_row_hash"
        ")"
    ), item)
        return True, result.lastrowid
    except IntegrityError:
        row = db.execute(text(
            "SELECT id FROM alerts WHERE content_hash = :content_hash OR unique_hash = :unique_hash LIMIT 1"
        ), {
            "content_hash": item["content_hash"],
            "unique_hash": item["unique_hash"],
        }).fetchone()
        return False, row[0] if row else None


def _process_excel(file_path: str, source_file: str, import_id: int, db) -> dict:
    """Process Excel in streaming mode — row-by-row, batch commits, bounded memory."""
    import openpyxl

    imported_at = datetime.now().isoformat()
    totals = {
        "inserted": 0,
        "skipped": 0,
        "failed": 0,
        "raw_only": 0,
        "total_rows": 0,
        "parsed_rows": 0,
        "failures": [],
    }

    # Speed up SQLite for bulk import
    db.execute(text("PRAGMA synchronous=NORMAL"))
    db.execute(text("PRAGMA cache_size=-8000"))
    db.execute(text("PRAGMA temp_store=MEMORY"))

    # Dedup within the current workbook only. Cross-workbook dedup is handled
    # by the per-batch DB query in _flush_batch, avoiding a blocking full-table scan.
    workbook_seen_hashes: set[str] = set()

    BATCH = 500

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            ws = workbook[sheet_name]

            # Read header row
            header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
            if header_cells is None:
                continue
            headers = [str(c.value) if c.value is not None else "" for c in header_cells]
            columns = _resolve_columns(headers)

            sheet_started_at = datetime.now().isoformat()
            sheet_id = db.execute(text(
                "INSERT INTO import_sheets ("
                "import_id, sheet_name, sheet_index, header_row, headers_json, row_count, "
                "parsed_rows, raw_rows, failed_rows, status, created_at"
                ") VALUES (:import_id, :sheet_name, :sheet_index, 1, :headers_json, 0, 0, 0, 0, 'processing', :created_at)"
            ), {
                "import_id": import_id,
                "sheet_name": sheet_name,
                "sheet_index": sheet_index,
                "headers_json": json.dumps(headers, ensure_ascii=False),
                "created_at": sheet_started_at,
            }).lastrowid

            sheet_counts = {"row_count": 0, "parsed": 0, "raw": 0, "failed": 0}
            batch_counter = 0  # rows already flushed for this sheet

            # Buffers for batch processing
            raw_row_buffer = []
            alert_buffer = []
            update_buffer = []
            row_counter = 0

            for row_tuple in ws.iter_rows(min_row=2, values_only=True):
                row_counter += 1
                excel_row_number = row_counter + 1
                totals["total_rows"] += 1
                sheet_counts["row_count"] += 1

                # Build row dict from cell values
                row_dict = {}
                for ci, cell_val in enumerate(row_tuple):
                    if ci < len(headers):
                        row_dict[headers[ci]] = cell_val

                raw_payload = {str(k): _json_safe(v) for k, v in row_dict.items()}
                raw_json = json.dumps(raw_payload, ensure_ascii=False, default=str)
                raw_hash = _row_hash(source_file, sheet_name, excel_row_number, raw_json)

                raw_row_buffer.append({
                    "import_id": import_id,
                    "sheet_id": sheet_id,
                    "source_file": source_file,
                    "sheet_name": sheet_name,
                    "excel_row_number": excel_row_number,
                    "raw_json": raw_json,
                    "row_hash": raw_hash,
                    "created_at": imported_at,
                })

                # Parse the row
                try:
                    item, parse_error = _alert_from_row(row_dict, columns)
                    if parse_error:
                        sheet_counts["raw"] += 1
                        totals["raw_only"] += 1
                        update_buffer.append({
                            "parse_status": "raw_only",
                            "parse_error": parse_error,
                            "normalized_json": None,
                            "alert_id": None,
                        })
                    else:
                        content_hash = item["content_hash"]
                        if content_hash in workbook_seen_hashes:
                            totals["skipped"] += 1
                            sheet_counts["parsed"] += 1
                            totals["parsed_rows"] += 1
                            update_buffer.append({
                                "parse_status": "skipped_duplicate",
                                "parse_error": None,
                                "normalized_json": json.dumps(item, ensure_ascii=False, default=str),
                                "alert_id": None,
                            })
                        else:
                            workbook_seen_hashes.add(content_hash)
                            totals["inserted"] += 1
                            totals["parsed_rows"] += 1
                            sheet_counts["parsed"] += 1

                            item.update({
                                "source_file": source_file,
                                "imported_at": imported_at,
                                "import_id": import_id,
                                "import_sheet_id": sheet_id,
                                "import_row_id": None,
                                "sheet_name": sheet_name,
                                "excel_row_number": excel_row_number,
                                "raw_row_hash": raw_hash,
                            })
                            alert_buffer.append(item)
                            update_buffer.append({
                                "parse_status": "parsed",
                                "parse_error": None,
                                "normalized_json": json.dumps(item, ensure_ascii=False, default=str),
                                "alert_id": None,
                                "content_hash": content_hash,
                            })

                except Exception as e:
                    sheet_counts["failed"] += 1
                    totals["failed"] += 1
                    totals["failures"].append({
                        "sheet": sheet_name,
                        "row": excel_row_number,
                        "error": str(e),
                    })
                    update_buffer.append({
                        "parse_status": "failed",
                        "parse_error": str(e),
                        "normalized_json": None,
                        "alert_id": None,
                        "content_hash": None,
                    })

                # Batch flush when buffer reaches BATCH size
                if len(raw_row_buffer) >= BATCH:
                    batch_counter = _flush_batch(
                        db, import_id, sheet_id, sheet_counts, totals,
                        raw_row_buffer, alert_buffer, update_buffer,
                        workbook_seen_hashes, source_file, sheet_name,
                        batch_counter,
                    )
                    raw_row_buffer.clear()
                    alert_buffer.clear()
                    update_buffer.clear()

            # Flush remaining rows
            if raw_row_buffer:
                batch_counter = _flush_batch(
                    db, import_id, sheet_id, sheet_counts, totals,
                    raw_row_buffer, alert_buffer, update_buffer,
                    workbook_seen_hashes, source_file, sheet_name,
                    batch_counter,
                )

            # Final sheet update
            db.execute(text(
                "UPDATE import_sheets SET row_count = :row_count, parsed_rows = :parsed, "
                "raw_rows = :raw, failed_rows = :failed, status = :status WHERE id = :id"
            ), {
                "row_count": sheet_counts["row_count"],
                "parsed": sheet_counts["parsed"],
                "raw": sheet_counts["raw"],
                "failed": sheet_counts["failed"],
                "status": "success" if sheet_counts["failed"] == 0 and sheet_counts["raw"] == 0 else "partial",
                "id": sheet_id,
            })
            _update_import(
                db, import_id,
                rows_inserted=totals["inserted"],
                rows_skipped=totals["skipped"],
                rows_failed=totals["failed"],
                total_rows=totals["total_rows"],
                parsed_rows=totals["parsed_rows"],
                raw_rows=totals["raw_only"],
                status="processing",
            )
            db.commit()

    except Exception as e:
        totals["failed"] += 1
        totals["failures"].append({"sheet": sheet_name, "row": None, "error": str(e)})
        try:
            db.execute(text(
                "UPDATE import_sheets SET failed_rows = 1, status = 'failed' WHERE id = :id"
            ), {"id": sheet_id})
            db.commit()
        except Exception:
            db.rollback()
    finally:
        workbook.close()

    # Restore SQLite settings
    db.execute(text("PRAGMA synchronous=FULL"))
    db.commit()
    return totals


def _flush_batch(db, import_id, sheet_id, sheet_counts, totals,
                 raw_row_buffer, alert_buffer, update_buffer,
                 workbook_seen_hashes, source_file, sheet_name,
                 batch_counter):
    """Insert one batch of rows into DB, parse alerts, update import_rows.

    batch_counter: running count of rows already inserted for this sheet.
    Returns the updated batch_counter.
    """
    # Step 1: Batch insert import_rows
    placeholders = []
    values = {}
    for bi, p in enumerate(raw_row_buffer):
        placeholders.append(
            f"(:imp_{bi}, :sh_{bi}, :sf_{bi}, :sn_{bi}, "
            f":ern_{bi}, :rj_{bi}, 'uploaded', NULL, :rh_{bi}, NULL, :ca_{bi})"
        )
        values.update({
            f"imp_{bi}": p["import_id"],
            f"sh_{bi}": p["sheet_id"],
            f"sf_{bi}": p["source_file"],
            f"sn_{bi}": p["sheet_name"],
            f"ern_{bi}": p["excel_row_number"],
            f"rj_{bi}": p["raw_json"],
            f"rh_{bi}": p["row_hash"],
            f"ca_{bi}": p["created_at"],
        })
    db.execute(text(
        f"INSERT INTO import_rows ("
        f"import_id, import_sheet_id, source_file, sheet_name, excel_row_number, "
        f"raw_json, parse_status, parse_error, row_hash, alert_id, created_at"
        f") VALUES {', '.join(placeholders)}"
    ), values)
    db.commit()

    new_counter = batch_counter + len(raw_row_buffer)

    # Step 2: Cross-workbook dedup — check which content_hashes in this batch
    # already exist in the database. Only query the small set of hashes in
    # this batch, avoiding the blocking full-table scan from the old code.
    db_existing = set()
    hash_to_alert_id = {}
    if alert_buffer:
        batch_hashes = [item["content_hash"] for item in alert_buffer]
        ph = ", ".join(f":h_{i}" for i in range(len(batch_hashes)))
        hp = {f"h_{i}": h for i, h in enumerate(batch_hashes)}
        existing = db.execute(text(
            f"SELECT content_hash FROM alerts WHERE content_hash IN ({ph})"
        ), hp).fetchall()
        db_existing = {r[0] for r in existing}

    # Step 3: Filter out DB duplicates from the alert insert; mark them as skipped.
    filtered_alerts = []
    db_skip_indices = set()  # indices into alert_buffer / update_buffer that are DB dupes
    for i, item in enumerate(alert_buffer):
        ch = item["content_hash"]
        if ch in db_existing:
            db_skip_indices.add(i)
            totals["skipped"] += 1
            totals["parsed_rows"] -= 1
            sheet_counts["parsed"] -= 1
            # Corresponding update_buffer entry is at the same index as alert_buffer
            # but update_buffer also has entries for raw_only/failed rows.
            # We need to find the correct index — see Step 4 below.
        else:
            filtered_alerts.append(item)

    # Step 4: Batch insert alerts (only non-duplicates)
    if filtered_alerts:
        cols = [
            "device_id", "first_alert_time", "last_alert_time", "source_ip", "target",
            "target_type", "port", "threat_type", "threat_level", "std_apt_org",
            "apt_org", "apt_org_tier", "alert_count", "vendors", "protocol",
            "intel_tags", "dns_resolved_ip", "down_traffic", "up_traffic", "asset_type",
            "source_file", "imported_at", "unique_hash", "content_hash", "import_id",
            "import_sheet_id", "import_row_id", "sheet_name", "excel_row_number", "raw_row_hash",
        ]
        alert_placeholders = []
        alert_values = {}
        for bi, item in enumerate(filtered_alerts):
            named = [f":{col}_{bi}" for col in cols]
            alert_placeholders.append(f"({', '.join(named)})")
            for col in cols:
                alert_values[f"{col}_{bi}"] = item.get(col)
        db.execute(text(
            f"INSERT INTO alerts ({', '.join(cols)}) VALUES {', '.join(alert_placeholders)}"
        ), alert_values)

        # Fetch back alert IDs by content_hash
        hash_to_alert_id = _load_alert_id_map_for_hashes(
            db, [item["content_hash"] for item in filtered_alerts]
        )

    # Step 5: Update import_rows with parse results and alert_ids.
    # Use row_hash for matching instead of unreliable offset-based ID lookup.
    # Build a map of row_hash from raw_row_buffer for quick access.
    hash_map = {p["row_hash"] for p in raw_row_buffer}

    # Track which update_buffer entry corresponds to which alert_buffer entry.
    # Both buffers are appended in the same order in the main loop: for each row
    # processed, exactly one entry goes into each buffer. But update_buffer also
    # has entries for raw_only/failed rows that don't have a corresponding alert.
    # So we need to match by position: the i-th "parsed" entry in update_buffer
    # corresponds to the i-th entry in alert_buffer.
    alert_buf_idx = 0
    for ui, upd in enumerate(update_buffer):
        upd_clean = dict(upd)
        upd_clean.pop("content_hash", None)

        if upd["parse_status"] == "parsed":
            # This update_buffer entry corresponds to alert_buffer[alert_buf_idx]
            row_h = raw_row_buffer[ui]["row_hash"] if ui < len(raw_row_buffer) else None

            if alert_buf_idx in db_skip_indices:
                # This row is a DB duplicate — mark as skipped, don't assign alert_id
                upd_clean["parse_status"] = "skipped_duplicate"
                upd_clean["alert_id"] = None
                upd_clean["normalized_json"] = upd.get("normalized_json")
            elif row_h and row_h in hash_map:
                ch = upd.get("content_hash")
                alert_id = hash_to_alert_id.get(ch) if ch else None
                upd_clean["alert_id"] = alert_id
                db.execute(text(
                    "UPDATE import_rows SET parse_status = :parse_status, parse_error = :parse_error, "
                    "normalized_json = :normalized_json, alert_id = :alert_id "
                    "WHERE row_hash = :row_hash AND import_id = :import_id"
                ), {**upd_clean, "row_hash": row_h, "import_id": import_id})
                # Also update alert_buffer's import_row_id for downstream use
                if alert_buf_idx < len(alert_buffer) and alert_buf_idx not in db_skip_indices:
                    alert_buffer[alert_buf_idx]["import_row_id"] = None  # row_id no longer tracked by id

            alert_buf_idx += 1
        else:
            # raw_only / failed — update by row_hash
            row_h = raw_row_buffer[ui]["row_hash"] if ui < len(raw_row_buffer) else None
            if row_h and row_h in hash_map:
                db.execute(text(
                    "UPDATE import_rows SET parse_status = :parse_status, parse_error = :parse_error, "
                    "normalized_json = :normalized_json, alert_id = :alert_id "
                    "WHERE row_hash = :row_hash AND import_id = :import_id"
                ), {**upd_clean, "row_hash": row_h, "import_id": import_id})

    # Progress update
    _update_import(
        db, import_id,
        raw_rows=totals["parsed_rows"] + totals["raw_only"],
        total_rows=totals["total_rows"],
        status="processing",
    )
    db.commit()

    return new_counter


def _run_import_job(import_id: int, file_path: str, source_file: str):
    logger.info("Starting import job: id=%d, file=%s", import_id, source_file)
    db = get_session_local()()
    try:
        _update_import(db, import_id, status="processing", queue_position=None)
        db.commit()

        result = _process_excel(file_path, source_file, import_id, db)
        status = "success"
        if result["failed"] > 0 or result["raw_only"] > 0:
            status = "partial" if (result["inserted"] or result["skipped"] or result["raw_only"]) else "failed"

        _update_import(
            db,
            import_id,
            rows_inserted=result["inserted"],
            rows_skipped=result["skipped"],
            rows_failed=result["failed"],
            total_rows=result["total_rows"],
            parsed_rows=result["parsed_rows"],
            raw_rows=result["raw_only"],
            status=status,
            log=json.dumps(result["failures"], ensure_ascii=False),
        )
        write_audit(db, "import_excel", "import", import_id, {
            "source_file": source_file,
            "inserted": result["inserted"],
            "skipped": result["skipped"],
            "raw_only": result["raw_only"],
            "failed": result["failed"],
            "total_rows": result["total_rows"],
        })
        db.commit()
        if status in {"success", "partial"}:
            logger.info("Import job succeeded: id=%d, inserted=%d, skipped=%d", import_id, result["inserted"], result["skipped"])
            # Snapshot rebuild: with streaming import, data is already in alerts table.
            # The snapshot is built on-demand when /api/alert-candidates is queried,
            # so we don't trigger a blocking rebuild here. The cache layer in alerts.py
            # will handle lazy construction on first query after import.
    except Exception as e:
        logger.exception("Import job failed: id=%d, file=%s", import_id, source_file)
        _update_import(
            db,
            import_id,
            rows_inserted=0,
            rows_skipped=0,
            rows_failed=1,
            total_rows=0,
            parsed_rows=0,
            raw_rows=0,
            status="failed",
            log=json.dumps([{"row": None, "error": str(e)}], ensure_ascii=False),
        )
        write_audit(db, "import_excel_failed", "import", import_id, {
            "source_file": source_file,
            "error": str(e),
        })
        db.commit()
    finally:
        db.close()


def _import_queue_worker():
    """Background worker: processes import queue one file at a time."""
    logger.info("Import queue worker started")
    while True:
        try:
            import_id, file_path, source_file = _import_queue.get()
            if import_id is None:
                logger.info("Import queue worker: received sentinel, stopping")
                break  # sentinel to stop worker
            logger.info("Import queue worker: processing import_id=%d, file=%s", import_id, source_file)
            _run_import_job(import_id, file_path, source_file)
            logger.info("Import queue worker: finished import_id=%d", import_id)
            _import_queue.task_done()
        except Exception:
            logger.exception("Import queue worker: unexpected error in worker loop")
            _import_queue.task_done()


def _ensure_queue_worker():
    """Start the background queue worker thread if not already running."""
    global _queue_worker_started
    if not _queue_worker_started:
        _queue_worker_started = True
        t = threading.Thread(target=_import_queue_worker, daemon=True, name="import-queue-worker")
        t.start()


def _repair_import_rows_if_needed(db, import_id):
    pending = db.execute(text(
        "SELECT COUNT(*) FROM import_rows WHERE import_id = :id AND parse_status = 'uploaded'"
    ), {"id": import_id}).scalar() or 0
    if pending == 0:
        return

    rows = db.execute(text(
        "SELECT id, raw_json, import_sheet_id, excel_row_number "
        "FROM import_rows WHERE import_id = :id "
        "ORDER BY import_sheet_id, excel_row_number"
    ), {"id": import_id}).fetchall()
    parsed_rows = db.execute(text(
        "SELECT import_row_id, id AS alert_id FROM alerts "
        "WHERE import_id = :id AND import_row_id IS NOT NULL"
    ), {"id": import_id}).fetchall()
    parsed_alert_map = {row[0]: {"alert_id": row[1]} for row in parsed_rows}

    updates, _stats = _rebuild_import_row_updates_for_repair(
        [dict(row._mapping) for row in rows],
        parsed_alert_map=parsed_alert_map,
        existing_hashes=set(),
    )
    for upd in updates:
        db.execute(text(
            "UPDATE import_rows SET parse_status = :parse_status, parse_error = :parse_error, "
            "normalized_json = :normalized_json, alert_id = COALESCE(:alert_id, alert_id) WHERE id = :id"
        ), upd)
    db.commit()


async def _save_upload_file(file: UploadFile, file_path: str) -> str:
    """Save uploaded file to disk while computing SHA256 hash. Returns the hex digest."""
    h = hashlib.sha256()
    with open(file_path, "wb") as handle:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
            handle.write(chunk)
    await file.close()
    return h.hexdigest()


def _start_import_thread(import_id: int, file_path: str, source_file: str):
    """Enqueue an import job into the queue for sequential processing."""
    _ensure_queue_worker()
    _import_queue.put((import_id, file_path, source_file))
    # Update queue position for all queued imports
    _update_queue_positions()


def _update_queue_positions():
    """Update queue_position field for all queued imports."""
    try:
        db = get_session_local()()
        position = 1
        for item in list(_import_queue.queue):
            if item[0] is not None:  # skip sentinel
                db.execute(text(
                    "UPDATE imports SET queue_position = :pos WHERE id = :id AND status = 'queued'"
                ), {"pos": position, "id": item[0]})
                position += 1
        db.commit()
    except Exception:
        pass
    finally:
        try:
            db.close()
        except Exception:
            pass


@router.post("")
async def import_excel(files: list[UploadFile] = File(...), db=Depends(get_db)):
    upload_dir = get_path("upload_tmp")
    os.makedirs(upload_dir, exist_ok=True)
    jobs = []
    session_local = get_session_local()

    for file in files:
        filename = os.path.basename(file.filename)
        now = datetime.now().isoformat()

        # Step 1: Save file and compute SHA256 hash
        stored_name = f"import_tmp_{filename}"
        file_path = os.path.join(upload_dir, stored_name)
        file_hash = await _save_upload_file(file, file_path)

        # Step 2: Check if same file was already imported (by file hash)
        session_db = session_local()
        try:
            existing = session_db.execute(text(
                "SELECT id, source_file, status FROM imports WHERE file_hash = :hash ORDER BY imported_at DESC LIMIT 1"
            ), {"hash": file_hash}).fetchone()

            if existing:
                existing_id, existing_file, existing_status = existing
                # File already uploaded — handle based on existing status
                if existing_status in ("success", "partial"):
                    # Already imported successfully, skip
                    os.remove(file_path)
                    jobs.append({
                        "id": existing_id,
                        "source_file": existing_file,
                        "status": existing_status,
                        "duplicate": True,
                    })
                    continue
                elif existing_status in ("queued", "processing", "uploaded"):
                    # Still being processed — don't create duplicate record,
                    # but also don't skip. Return existing record so frontend can poll.
                    os.remove(file_path)
                    jobs.append({
                        "id": existing_id,
                        "source_file": existing_file,
                        "status": existing_status,
                        "duplicate": True,
                    })
                    continue
                elif existing_status == "failed":
                    # Previous import failed — re-enqueue for processing
                    logger.info("Re-enqueueing failed import: id=%d, file=%s", existing_id, existing_file)
                    # Update status back to queued and continue to enqueue
                    session_db.execute(text(
                        "UPDATE imports SET status = 'queued', queue_position = NULL WHERE id = :id"
                    ), {"id": existing_id})
                    session_db.commit()
                    import_id = existing_id
                    # Rename temp file and enqueue
                    final_name = f"import_{import_id}_{filename}"
                    final_path = os.path.join(upload_dir, final_name)
                    os.rename(file_path, final_path)
                    _start_import_thread(import_id, final_path, filename)
                    jobs.append({"id": import_id, "source_file": filename, "status": "queued"})
                    continue
                else:
                    # Unknown status — treat as new upload
                    pass
        finally:
            session_db.close()

        # Step 3: Create import record with file_hash, status=queued
        holder = {}
        record_db = session_local()

        def _create_import_record():
            cursor = record_db.execute(text(
                "INSERT INTO imports ("
                "source_file, imported_at, rows_inserted, rows_skipped, rows_failed, "
                "total_rows, parsed_rows, raw_rows, status, log, file_hash"
                ") VALUES (:source_file, :imported_at, 0, 0, 0, 0, 0, 0, 'queued', '[]', :file_hash)"
            ), {"source_file": filename, "imported_at": now, "file_hash": file_hash})
            holder["import_id"] = cursor.lastrowid
            write_audit(record_db, "upload_import_file", "import", holder["import_id"], {"source_file": filename})
            record_db.commit()

        try:
            _run_locked_retry(record_db, _create_import_record)
        finally:
            record_db.close()

        import_id = holder["import_id"]
        # Rename the temp file to match the import_id
        final_name = f"import_{import_id}_{filename}"
        final_path = os.path.join(upload_dir, final_name)
        os.rename(file_path, final_path)

        # Step 4: Enqueue for sequential processing
        _start_import_thread(import_id, final_path, filename)
        jobs.append({"id": import_id, "source_file": filename, "status": "queued"})

    return {"accepted": len(jobs), "jobs": jobs}


def _import_item(row):
    d = dict(row._mapping)
    if d.get("imported_at"):
        d["imported_at"] = str(d["imported_at"])
    for field in ("rows_inserted", "rows_skipped", "rows_failed", "total_rows", "parsed_rows", "raw_rows"):
        d[field] = d.get(field) or 0
    if d.get("queue_position") is not None:
        d["queue_position"] = int(d["queue_position"])
    d["scope"] = {
        "imports_all_sheets": True,
        "imports_original_rows": True,
        "dedupe_policy": "exact_alert_content_with_time",
        "traceability": "alert.import_row_id -> import_rows.id",
    }
    return d


@router.get("")
def list_imports(db=Depends(get_db)):
    rows = db.execute(text("SELECT * FROM imports ORDER BY imported_at DESC")).fetchall()
    return [_import_item(row) for row in rows]


@router.post("/reprocess-queued")
def reprocess_queued_imports(db=Depends(get_db)):
    """Re-enqueue all imports stuck in 'queued' status.

    This is useful after backend restart when the in-memory queue was lost.
    """
    stuck = db.execute(text(
        "SELECT id, source_file FROM imports WHERE status = 'queued' ORDER BY imported_at ASC"
    )).fetchall()
    if not stuck:
        return {"ok": True, "requeued": 0, "message": "没有卡住的导入任务"}

    upload_dir = get_path("upload_tmp")
    requeued = []
    for row in stuck:
        imp_id, source_file = row
        # Check if the uploaded file still exists
        expected_file = os.path.join(upload_dir, f"import_{imp_id}_{source_file}")
        if not os.path.isfile(expected_file):
            logger.warning("Reprocess: file not found for import_id=%d, marking as failed", imp_id)
            db.execute(text(
                "UPDATE imports SET status = 'failed', log = :log WHERE id = :id"
            ), {"id": imp_id, "log": json.dumps([{"error": "上传文件已丢失，无法重新处理"}])})
            db.commit()
            continue

        # Reset status and re-enqueue
        db.execute(text(
            "UPDATE imports SET status = 'queued', queue_position = NULL WHERE id = :id"
        ), {"id": imp_id})
        db.commit()
        _start_import_thread(imp_id, expected_file, source_file)
        requeued.append(imp_id)
        logger.info("Re-queued stuck import: id=%d, file=%s", imp_id, source_file)

    return {"ok": True, "requeued": len(requeued), "ids": requeued}


@router.get("/{import_id}")
def get_import(import_id: int, db=Depends(get_db)):
    row = db.execute(text("SELECT * FROM imports WHERE id = :id"), {"id": import_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Import not found")
    return _import_item(row)


@router.get("/{import_id}/sheets")
def list_import_sheets(import_id: int, db=Depends(get_db)):
    rows = db.execute(text(
        "SELECT * FROM import_sheets WHERE import_id = :import_id ORDER BY sheet_index"
    ), {"import_id": import_id}).fetchall()
    items = []
    for row in rows:
        item = dict(row._mapping)
        item["headers"] = json.loads(item.get("headers_json") or "[]")
        if item.get("created_at"):
            item["created_at"] = str(item["created_at"])
        items.append(item)
    return items


@router.get("/{import_id}/rows")
def list_import_rows(
    import_id: int,
    sheet_id: int = Query(None),
    status: str = Query(None),
    status_group: str = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
    db=Depends(get_db),
):
    _repair_import_rows_if_needed(db, import_id)
    where = ["import_id = :import_id"]
    params = {"import_id": import_id}
    if sheet_id:
        where.append("import_sheet_id = :sheet_id")
        params["sheet_id"] = sheet_id
    statuses = _resolve_row_status_filters(status=status, status_group=status_group)
    if status_group and not statuses:
        raise HTTPException(status_code=400, detail="Invalid status_group")
    if statuses:
        if len(statuses) == 1:
            where.append("parse_status = :status")
            params["status"] = statuses[0]
        else:
            placeholders = []
            for idx, item in enumerate(statuses):
                key = f"status_{idx}"
                placeholders.append(f":{key}")
                params[key] = item
            where.append(f"parse_status IN ({', '.join(placeholders)})")
    where_sql = " AND ".join(where)
    total = db.execute(text(f"SELECT COUNT(*) FROM import_rows WHERE {where_sql}"), params).scalar() or 0
    params["limit"] = page_size
    params["offset"] = (page - 1) * page_size
    rows = db.execute(text(f"""
        SELECT * FROM import_rows
        WHERE {where_sql}
        ORDER BY import_sheet_id, excel_row_number
        LIMIT :limit OFFSET :offset
    """), params).fetchall()

    items = []
    for row in rows:
        item = dict(row._mapping)
        item["raw"] = json.loads(item.get("raw_json") or "{}")
        item["normalized"] = json.loads(item.get("normalized_json") or "{}") if item.get("normalized_json") else None
        if item.get("created_at"):
            item["created_at"] = str(item["created_at"])
        items.append(item)

    if total == 0 and status_group == "issues":
        import_row = db.execute(text(
            "SELECT log FROM imports WHERE id = :id"
        ), {"id": import_id}).fetchone()
        failures = json.loads(import_row[0] or "[]") if import_row else []
        synthetic_items = []
        for idx, failure in enumerate(failures):
            synthetic_items.append({
                "id": f"log-{idx}",
                "import_id": import_id,
                "import_sheet_id": None,
                "source_file": None,
                "sheet_name": failure.get("sheet"),
                "excel_row_number": failure.get("row"),
                "raw": {},
                "normalized": None,
                "parse_status": "failed",
                "parse_error": failure.get("error"),
                "alert_id": None,
                "created_at": None,
            })
        if synthetic_items:
            total = len(synthetic_items)
            start = (page - 1) * page_size
            end = start + page_size
            items = synthetic_items[start:end]

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/{import_id}/failures.csv")
def download_failures(
    import_id: int,
    type: str = Query("failures"),
    db=Depends(get_db),
):
    statuses = CSV_EXPORT_TYPES.get(type)
    if not statuses:
        raise HTTPException(status_code=400, detail="Invalid export type")
    status_placeholders = ", ".join(f":status_{idx}" for idx in range(len(statuses)))
    params = {"id": import_id, **{f"status_{idx}": item for idx, item in enumerate(statuses)}}
    rows = db.execute(text(
        "SELECT sheet_name, excel_row_number, parse_status, parse_error, raw_json, normalized_json "
        "FROM import_rows "
        f"WHERE import_id = :id AND parse_status IN ({status_placeholders}) "
        "ORDER BY import_sheet_id, excel_row_number"
    ), params).fetchall()
    if not rows:
        row = db.execute(text(
            "SELECT source_file, log FROM imports WHERE id = :id"
        ), {"id": import_id}).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Import not found")
        if type != "failures":
            failures = []
        else:
            failures = json.loads(row[1] or "[]")
    else:
        failures = _build_failure_export_rows([
            {
                "sheet_name": row[0],
                "excel_row_number": row[1],
                "parse_status": row[2],
                "parse_error": row[3],
                "raw_json": row[4],
                "normalized_json": row[5],
            }
            for row in rows
        ])
    if not failures:
        raise HTTPException(status_code=404, detail="No failures for this import")

    buf = io.StringIO()
    fieldnames = sorted({key for item in failures for key in item.keys()}) or ["error"]
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    for item in failures:
        writer.writerow(item)
    data = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    filename = f"import_{import_id}_{type}.csv"
    return StreamingResponse(
        data,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.delete("/all")
def delete_all_imports(backup: str = Query("false"), db=Depends(get_db)):
    """删除所有已上传的 Excel 数据（告警、导入记录、明细等）。

    当 backup="true" 时，先备份当前 alerts 表到 backups/ 目录。
    """
    import shutil as _shutil

    backup_lower = str(backup).lower()
    do_backup = backup_lower in ("true", "1", "yes")

    # Count what we're about to delete
    alert_count = db.execute(text("SELECT COUNT(*) FROM alerts")).scalar() or 0
    import_count = db.execute(text("SELECT COUNT(*) FROM imports")).scalar() or 0
    row_count = db.execute(text("SELECT COUNT(*) FROM import_rows")).scalar() or 0
    sheet_count = db.execute(text("SELECT COUNT(*) FROM import_sheets")).scalar() or 0

    if alert_count == 0 and import_count == 0:
        return {"ok": True, "message": "没有数据可删除", "deleted": {"alerts": 0, "imports": 0, "rows": 0, "sheets": 0}}

    backup_path = ""
    if do_backup:
        db_path = get_path("db")
        backups_dir = SCRIPT_DIR / "backups"
        backups_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backups_dir / f"workbench_before_clear_{timestamp}.db"
        try:
            _shutil.copy2(db_path, backup_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"数据库备份失败: {e}")

    # Delete in order: rows -> sheets -> imports -> alerts (FK safety)
    db.execute(text("DELETE FROM import_rows"))
    db.execute(text("DELETE FROM import_sheets"))
    db.execute(text("DELETE FROM imports"))
    db.execute(text("DELETE FROM alerts"))
    write_audit(db, "delete_all_imports", "import", None, {
        "alerts_deleted": alert_count,
        "imports_deleted": import_count,
        "backup": backup_path or None,
    })
    db.commit()
    rebuild_candidate_snapshots(db)

    # WAL checkpoint
    db.close()
    engine = get_engine()
    raw_conn = engine.raw_connection()
    try:
        _checkpoint_after_import_delete(raw_conn)
    except Exception:
        pass
    finally:
        if raw_conn is not None:
            raw_conn.close()

    return {
        "ok": True,
        "message": "全部数据已删除",
        "deleted": {
            "alerts": alert_count,
            "imports": import_count,
            "rows": row_count,
            "sheets": sheet_count,
        },
        "backup": backup_path or None,
    }


@router.delete("/{import_id}")
def delete_import(import_id: int, db=Depends(get_db)):
    row = db.execute(text("SELECT source_file FROM imports WHERE id = :id"), {"id": import_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Import not found")
    source_file = row[0]
    db.execute(text("DELETE FROM alerts WHERE import_id = :import_id"), {"import_id": import_id})
    db.execute(text(
        "DELETE FROM alerts WHERE source_file = :source_file AND import_id IS NULL"
    ), {"source_file": source_file})
    db.execute(text("DELETE FROM import_rows WHERE import_id = :id"), {"id": import_id})
    db.execute(text("DELETE FROM import_sheets WHERE import_id = :id"), {"id": import_id})
    db.execute(text("DELETE FROM imports WHERE id = :id"), {"id": import_id})
    write_audit(db, "delete_import", "import", import_id, {"source_file": source_file})
    db.commit()
    rebuild_candidate_snapshots(db)

    # Delete the uploaded Excel file from uploads/
    upload_dir = get_path("upload_tmp")
    stored_name = f"import_{import_id}_{source_file}"
    file_path = os.path.join(upload_dir, stored_name)
    if os.path.isfile(file_path):
        try:
            os.remove(file_path)
        except OSError:
            pass

    # Close the ORM session before VACUUM to release any implicit locks
    db.close()

    # Keep post-delete cleanup lightweight. We only checkpoint WAL here and
    # avoid proactively compacting the main DB file on every import delete.
    engine = get_engine()
    raw_conn = engine.raw_connection()
    try:
        _checkpoint_after_import_delete(raw_conn)
    except Exception:
        pass
    finally:
        if raw_conn is not None:
            raw_conn.close()

    return {"ok": True}


@router.post("/{import_id}/repair-metadata")
def repair_import_metadata(import_id: int, db=Depends(get_db)):
    """Re-parse import_rows.raw_json to backfill missing metadata fields (analysis_status, is_focused).

    This avoids re-uploading Excel files. The raw_json already contains the original
    Excel cell values; we just need to re-extract them using the current parsing logic.
    """
    # Verify import exists
    row = db.execute(text("SELECT id FROM imports WHERE id = :id"), {"id": import_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Import not found")

    # Fetch all parsed import_rows that have an associated alert
    import_rows = db.execute(text(
        "SELECT ir.id, ir.raw_json, ir.alert_id, ir.normalized_json "
        "FROM import_rows ir "
        "WHERE ir.import_id = :id AND ir.parse_status = 'parsed' AND ir.alert_id IS NOT NULL"
    ), {"id": import_id}).fetchall()

    stats = {"total": len(import_rows), "repaired": 0, "skipped": 0, "errors": 0}

    for ir in import_rows:
        try:
            raw_payload = json.loads(ir[1] or "{}")
            columns = _resolve_columns(list(raw_payload.keys()))
            item, parse_error = _alert_from_row(raw_payload, columns)
            if parse_error or not item:
                stats["skipped"] += 1
                continue

            alert_id = ir[2]
            analysis_status = item.get("analysis_status")
            is_focused = item.get("is_focused", 0)

            # Update the alerts table
            db.execute(text(
                "UPDATE alerts SET analysis_status = :analysis_status, is_focused = :is_focused "
                "WHERE id = :alert_id"
            ), {
                "analysis_status": analysis_status,
                "is_focused": is_focused,
                "alert_id": alert_id,
            })

            # Update the normalized_json in import_rows
            new_normalized = json.dumps(item, ensure_ascii=False, default=str)
            if new_normalized != ir[3]:
                db.execute(text(
                    "UPDATE import_rows SET normalized_json = :normalized_json WHERE id = :id"
                ), {"normalized_json": new_normalized, "id": ir[0]})

            stats["repaired"] += 1
        except Exception:
            stats["errors"] += 1

    db.commit()
    rebuild_candidate_snapshots(db)
    return {"ok": True, "stats": stats}

